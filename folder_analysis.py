import os
import stat
import pandas as pd
import openpyxl
import logging
import time
from datetime import datetime
import win32com.client
import win32security

# Set output directory
OUTPUT_DIR = r"C:\projects\Genpact\folder_analyzer\Output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Setup timestamp for logger file name
timestamp = time.strftime("%Y%m%d_%H%M%S")

# Configure logging
logging.basicConfig(
    filename = os.path.join(OUTPUT_DIR, f'folder_analysis_{timestamp}.txt'),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class Folder_Analyzer:
    def __init__(self, input_path):
        self.input_path = input_path
        #Empty list to store folder data
        self.folder_data = [] 

    def format_size(self, size_in_bytes):
        """Format size to appropriate unit (KB, MB, GB) """
        kb = size_in_bytes / 1024.0
        mb = kb / 1024.0
        gb = mb / 1024.0

        if gb >= 1:
            return f"{round(gb, 2)} GB"
        elif mb >= 1:
            return f"{round(mb, 2)} MB"
        else:
            return f"{round(kb, 2)} KB"

    def get_last_modified_by(self, folder_path):
        """Get the user who last modified the folder."""
        try:
            folder = win32security.GetFileSecurity(folder_path, win32security.OWNER_SECURITY_INFORMATION)
            owner_sid = folder.GetSecurityDescriptorOwner()
            owner_name, domain, type = win32security.LookupAccountSid(None, owner_sid)
            return owner_name
        except Exception as e:
            logging.error(f"Error getting last modified by user for {folder_path}: {str(e)}")
            return "Unknown"

    def get_folder_stats(self, folder_path, depth=0):
        """ Get folder properties """
        total_size = 0
        all_folders_data = []
        file_count = 0
        #starts with -1 because root is included in os.walk
        subfolder_count = -1  

        try:
            #walk through directory
            for root, dirs, files in os.walk(folder_path):
                #Count all directories except root
                subfolder_count += 1

                #Count files and calculate size
                for file in files:
                    try:
                        #construct the full file path
                        full_file_path = os.path.join(root, file)
                        if os.path.isfile(full_file_path): #Verify it's a file
                            try:
                                file_stat = os.stat(full_file_path)
                                if file_stat.st_size > 0: #only counts if file has size
                                    total_size += file_stat.st_size
                                    file_count += 1
                            except (OSError, FileNotFoundError) as e:
                                logging.warning(f"Couldn't get size for {full_file_path}: {str(e)}")
                    except Exception as e:
                        logging.warning(f"Error accessing file {file} in {root}: {str(e)}")
                        continue
            
             # Get folder creation year
            creation_time = os.path.getctime(folder_path)
            creation_year = datetime.fromtimestamp(creation_time).year

             # Get folder last modified date
            modification_time = os.path.getmtime(folder_path)
            date_last_modified = datetime.fromtimestamp(modification_time).strftime('%Y-%m-%d')

            # Get last modified by user
            last_modified_by = self.get_last_modified_by(folder_path)

            #Format size with appropriate units
            size_formatted = self.format_size(total_size)

            #Collect folder properties data
            folder_data = {
                'depth': depth,
                'folder_name': os.path.basename(os.path.normpath(folder_path)) or folder_path,
                'folder_path': folder_path,
                'folder_size': size_formatted,
                'total_files': file_count,
                'total_subfolders': max(0, subfolder_count),
                'year_created': creation_year,
                'last_modified_date': date_last_modified,
                'last_modified_by': last_modified_by,  
                'folder_size_before_deletion': size_formatted,
                'folder_size_after_deletion': '',
                'scream_test': '',
                'renamed_date': '',
                'delete': '',
                'retain': '',
                'hold': '',
                'comment': ''
            }

            logging.info(f"Processed folder: {folder_path} | Depth: {depth} | Files: {file_count} | Subfolders: {subfolder_count} | Size: {size_formatted} | year_created: {creation_year} | last_modified_date: {date_last_modified} | Last Modified By: {last_modified_by}")
            all_folders_data.append(folder_data)

            # Recurse into subfolders
            with os.scandir(folder_path) as entries:
                for entry in entries:
                    if entry.is_dir():
                        subfolder_data = self.get_folder_stats(entry.path, depth + 1)
                        all_folders_data.extend(subfolder_data)

            return all_folders_data

        except Exception as e:
            logging.error(f"Error processing folder {folder_path}: {str(e)}")
            return []

    def create_excel(self, folder_data, output_file):
        try:
            wb = openpyxl.Workbook()

            property_headers = [
                'folder_name',
                'folder_path',
                'folder_size', 
                'total_files', 
                'total_subfolders', 
                'year_created',
                'last_modified_date',
                'last_modified_by', 
                'folder_size_before_deletion', 
                'folder_size_after_deletion',
                'scream_test', 
                'renamed_date', 
                'delete', 
                'retain',
                'hold', 
                'comment'
            ]

            #Sheet 1: Folder Hierarchy
            ws1 = wb.active
            ws1.title = "Folder_Hierarchy"

            max_depth = max(folder['depth'] for folder in folder_data)

            #Write level headers
            for i in range(max_depth + 1):
                ws1.cell(row=1, column=i + 1, value=f"Level {i}")

            root_name = os.path.basename(os.path.normpath(self.input_path)) or self.input_path
            base_path_parts = os.path.normpath(self.input_path).split(os.sep)

            # Fill in the rows
            for idx, folder in enumerate(folder_data, start=2):
                rel_path_parts = os.path.normpath(folder['folder_path']).split(os.sep)
                
                # Always start with root at Level 0
                ws1.cell(row=idx, column=1, value=root_name)

                # Then write subfolders starting from Level 1
                relative_parts = rel_path_parts[len(base_path_parts):]
                for level_idx, folder_part in enumerate(relative_parts):
                    ws1.cell(row=idx, column=level_idx + 2, value=folder_part)

            #Sheet 2: Folder Properties
            ws2 = wb.create_sheet(title="Folder_Properties")

            #Write property headers afer level column
            all_headers = ['depth'] + property_headers
            for col_idx, header in enumerate(all_headers, start=1):
                ws2.cell(row=1, column=col_idx, value=header)

            for row_idx, folder in enumerate(folder_data, start=2):
                for col_idx, header in enumerate(all_headers, start=1):
                    ws2.cell(row=row_idx, column=col_idx, value=folder.get(header, ''))

            wb.save(output_file)
            logging.info(f"Excel report created successfully: {output_file}")

        except Exception as e:
            logging.error(f"Failed to create Excel file: {str(e)}")

    def run_folder_analysis(self):
        base_path = self.input_path
        if not os.path.exists(base_path):
            logging.error("The specified path does not exist.")
            return

        logging.info(f"Starting hierarchical folder analysis for path: {base_path}")
        self.folder_data = self.get_folder_stats(base_path)

        if not self.folder_data:
            logging.warning("No folder data collected.")
            return

        output_file = os.path.join(OUTPUT_DIR, f'folder_hierarchy_analysis_{timestamp}.xlsx')
        self.create_excel(self.folder_data, output_file)

        df = pd.DataFrame(self.folder_data)
        total_bytes = df['folder_size'].apply(self.parse_size_to_bytes).sum()

        logging.info(f"Analysis complete. Report saved: {output_file}")
        logging.info(f"Total folders processed: {len(self.folder_data)}")
        logging.info(f"Total files: {df['total_files'].sum()}")
        logging.info(f"Total subfolders: {df['total_subfolders'].sum()}")
        logging.info(f"Total size: {self.format_size(total_bytes)}")

    def parse_size_to_bytes(self, size_str):
        """Helper to convert size strings back to bytes for aggregation."""
        try:
            value, unit = size_str.split()
            value = float(value)
            if unit == 'GB':
                return value * 1024 ** 3
            elif unit == 'MB':
                return value * 1024 ** 2
            elif unit == 'KB':
                return value * 1024
            else:
                return 0  
        except Exception as e:
            logging.error(f"Error parsing size {size_str}: {str(e)}")
            return 0

if __name__ == "__main__":
    INPUT_DIR = r"C:\projects"
    analyzer = Folder_Analyzer(INPUT_DIR)
    analyzer.run_folder_analysis()
